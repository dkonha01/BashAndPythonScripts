from openpyxl import *
import sys


wb = load_workbook(sys.argv[1])
sheet = wb.active

for row in sheet.iter_rows(7):
    for cell in row:
        if cell.value == sys.argv[2]:
            print(sheet.cell(row=cell.row, column=6).value)
