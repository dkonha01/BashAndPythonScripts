# hard coded example using openpyxl

from openpyxl import *
import sys
import re

# inputString = sys.argv[2]

wb = load_workbook(sys.argv[1])
sheet = wb.active

for row in sheet.iter_rows(6):
    for cell in row:
        cv = cell.value
        pattern = '^King\sTubby+'
        #   pattern = '^inputString'
        m = re.search(pattern, str(cv))
        if m:
            print(sheet.cell(row=cell.row, column=6).value)
